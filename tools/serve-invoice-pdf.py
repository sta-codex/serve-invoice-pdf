#!/usr/bin/env python
"""Small HTTP service for generating STA invoice and monthly-close documents.

Typical Airtable automation flow:
1. Airtable sends the invoice payload to POST /airtable/invoices/render.
2. This service renders a PDF and, for grouped invoices, a detail XLSX workbook.
3. Airtable receives temporary public URLs and attaches them to Documento.

Monthly-close flow:
1. Airtable sends contable existences to POST /airtable/monthly-close/render.
2. This service renders a summary plus one XLSX sheet per Packing List.
3. Airtable attaches the temporary URL to Packing Lists.Cierre copias.
"""

from __future__ import annotations

import base64
import importlib.util
import json
import os
import re
import secrets
import sys
import tempfile
import time
from datetime import date
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlparse

ROOT = Path(__file__).resolve().parents[1]
RENDER_SCRIPT = ROOT / "tools" / "render-sales-invoice-pdf.py"
REASSIGNMENT_MAP_HTML = ROOT / "tools" / "reassignment-map.html"
DEFAULT_HOST = "127.0.0.1"
DEFAULT_PORT = 8787
FILE_CACHE: dict[str, dict] = {}

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("Missing openpyxl. Install it with: python -m pip install openpyxl") from exc


def load_renderer():
    spec = importlib.util.spec_from_file_location("sta_invoice_renderer", RENDER_SCRIPT)
    if not spec or not spec.loader:
        raise RuntimeError(f"Cannot load renderer from {RENDER_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


RENDERER = load_renderer()
RENDERER.load_dotenv()

MONTHLY_CLOSE_FIELDS = (
    "ID Fábrica",
    "Neto",
    "Bruto",
    "Carta de crédito valor",
    "Material valor [€]",
    "Inspección valor",
    "Inspección origen valor",
    "Seguro valor",
    "Flete valor",
    "Comisiones de compra valor",
    "CBAM valor",
    "Desestiba valor",
    "T3 Valor",
    "Clasificación valor",
    "Entoldado valor",
    "Abono agencia valor",
    "DVD valor",
    "Despacho valor",
    "Plásticos valor",
    "Cuota valor",
    "Movimientos de compra base valor",
    "Movimientos de compra descarga valor",
    "Almacenajes valor",
    "Valor total",
    "Precio total",
)
MONTHLY_CLOSE_WEIGHT_FIELDS = {"Neto", "Bruto"}
MONTHLY_CLOSE_PRICE_FIELDS = {"Precio total"}
MONTHLY_CLOSE_NUMERIC_FIELDS = set(MONTHLY_CLOSE_FIELDS[1:])
MONTHLY_CLOSE_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def env_bool(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "si", "sí", "on"}


def service_config() -> dict[str, str | bool | int]:
    return {
        "base_id": os.environ.get("AIRTABLE_BASE_ID", RENDERER.DEFAULT_BASE_ID),
        "airtable_token": os.environ.get("AIRTABLE_TOKEN") or os.environ.get("AIRTABLE_PAT") or "",
        "attachment_field_id": os.environ.get("STA_INVOICE_ATTACHMENT_FIELD_ID", ""),
        "service_token": os.environ.get("STA_INVOICE_SERVICE_TOKEN", ""),
        "public_base_url": os.environ.get("STA_INVOICE_PUBLIC_BASE_URL", ""),
        "default_source": os.environ.get("STA_INVOICE_SOURCE", "live"),
        "default_upload": env_bool("STA_INVOICE_UPLOAD", True),
        "max_lines": int(os.environ.get("STA_INVOICE_MAX_LINES", "30")),
        "cache_ttl_seconds": int(os.environ.get("STA_INVOICE_CACHE_TTL_SECONDS", "900")),
    }


def json_response(handler: BaseHTTPRequestHandler, status: int, payload: dict) -> None:
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Allow-Headers", "Authorization, Content-Type, X-STA-Invoice-Token")
    handler.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def text_response(handler: BaseHTTPRequestHandler, status: int, body: str) -> None:
    data = body.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "text/plain; charset=utf-8")
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Allow-Headers", "Authorization, Content-Type, X-STA-Invoice-Token")
    handler.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def html_response(handler: BaseHTTPRequestHandler, status: int, body: bytes) -> None:
    handler.send_response(status)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Cache-Control", "public, max-age=300")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def read_json_body(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0") or "0")
    if not length:
        return {}
    data = handler.rfile.read(length)
    return json.loads(data.decode("utf-8"))


def is_authorized(handler: BaseHTTPRequestHandler, config: dict) -> bool:
    expected = str(config.get("service_token") or "")
    if not expected:
        return True
    bearer = handler.headers.get("Authorization", "")
    header_token = handler.headers.get("X-STA-Invoice-Token", "")
    return bearer == f"Bearer {expected}" or header_token == expected


def build_invoice(payload: dict, config: dict):
    source = str(payload.get("source") or config["default_source"])
    invoice_query = str(payload.get("recordId") or payload.get("invoice") or "").strip()
    if not invoice_query:
        raise ValueError("Missing recordId or invoice.")

    title = str(payload.get("title") or "Commercial Invoice Number")
    if payload.get("proforma"):
        title = "Proforma Invoice Number"
    category = str(payload.get("category") or "")
    vat_rate = payload.get("vatRate")
    vat_rate = float(vat_rate) if vat_rate is not None and vat_rate != "" else None

    if source == "local":
        invoice = RENDERER.build_invoice(
            records_dir=RENDERER.DEFAULT_RECORDS_DIR,
            invoice_query=invoice_query,
            title_label=title,
            category_override=category,
            vat_override=vat_rate,
        )
        return RENDERER.apply_invoice_mode(invoice, payload.get("mode"), invoice.weight_mode)

    token = str(config.get("airtable_token") or "")
    if not token:
        raise ValueError("Live mode requires AIRTABLE_TOKEN or AIRTABLE_PAT.")
    invoice = RENDERER.build_invoice_live(
        base_id=str(config["base_id"]),
        token=token,
        invoice_query=invoice_query,
        title_label=title,
        category_override=category,
        vat_override=vat_rate,
    )
    return RENDERER.apply_invoice_mode(invoice, payload.get("mode"), invoice.weight_mode)


def build_invoice_from_payload(payload: dict):
    title = str(payload.get("title") or "Commercial Invoice Number")
    if payload.get("proforma"):
        title = "Proforma Invoice Number"
    category = str(payload.get("category") or "")
    vat_rate = payload.get("vatRate")
    vat_rate = float(vat_rate) if vat_rate is not None and vat_rate != "" else None
    invoice_payload = payload.get("invoice") or payload
    if not isinstance(invoice_payload, dict):
        raise ValueError("Missing invoice payload.")
    return RENDERER.build_invoice_from_payload(
        invoice_payload,
        title_label=title,
        category_override=category,
        vat_override=vat_rate,
    )


def render_invoice_pdf_bytes(invoice, max_lines: int) -> tuple[bytes, str]:
    mode_suffix = "agrupada" if invoice.mode == "grouped" else "detalle"
    filename = f"{RENDERER.safe_filename(invoice.invoice_id)}-{mode_suffix}.pdf"
    with tempfile.TemporaryDirectory(prefix="sta-invoice-") as tmp_dir:
        path = Path(tmp_dir) / filename
        RENDERER.render_pdf(invoice, path, max_lines=max_lines)
        return path.read_bytes(), filename


def line_net_weight(line, weight_mode: str) -> float:
    if line.net_weight:
        return line.net_weight
    return line.quantity if weight_mode == "net" else 0.0


def line_gross_weight(line, weight_mode: str) -> float:
    if line.gross_weight:
        return line.gross_weight
    return line.quantity if weight_mode == "gross" else 0.0


def customer_uses_order_product(invoice) -> bool:
    return str(getattr(invoice, "customer_code", "") or "").strip() == "115"


def detail_headers(invoice, weight_mode: str) -> list[str]:
    if customer_uses_order_product(invoice):
        dynamic_headers = ["N\u00famero de orden", "C\u00f3digo de producto"]
    else:
        dynamic_headers = ["Acabado", "Color"]
    headers = [
        "item",
        "i",
        "Uds",
        "Material",
        "Espesor",
        "Ancho",
        "Largo",
        "Calidad",
        "Recubrimiento",
        *dynamic_headers,
        "ID F\u00e1brica",
    ]
    if weight_mode == "net":
        headers.extend(["Neto", "Bruto"])
    else:
        headers.append("Bruto")
    return headers


def line_extra_values(line, invoice) -> tuple[str, str]:
    if customer_uses_order_product(invoice):
        return line.order_number, line.product_code
    return line.finish, line.color


def line_detail_values(line, weight_mode: str, invoice, item_label: str, detail_index: int) -> list:
    extra_1, extra_2 = line_extra_values(line, invoice)
    values = [
        item_label,
        detail_index,
        line.units or None,
        line.material,
        RENDERER.fmt_thickness_text(line.thickness),
        RENDERER.fmt_measure_text(line.width),
        RENDERER.fmt_measure_text(line.length),
        line.quality,
        line.coating,
        extra_1,
        extra_2,
        line.factory_id,
    ]
    if weight_mode == "net":
        values.extend([line_net_weight(line, weight_mode), line_gross_weight(line, weight_mode)])
    else:
        values.append(line_gross_weight(line, weight_mode))
    return values


def line_group_values(group_line, weight_mode: str, invoice) -> list:
    item_label = group_line.item_label or group_line.factory_id
    extra_1, extra_2 = line_extra_values(group_line, invoice)
    base = [
        item_label,
        None,
        group_line.units or None,
        group_line.material,
        RENDERER.fmt_thickness_text(group_line.thickness),
        RENDERER.fmt_measure_text(group_line.width),
        RENDERER.fmt_measure_text(group_line.length),
        group_line.quality,
        group_line.coating,
        extra_1,
        extra_2,
        item_label,
    ]
    if weight_mode == "net":
        base.extend([line_net_weight(group_line, weight_mode), line_gross_weight(group_line, weight_mode)])
    else:
        base.append(line_gross_weight(group_line, weight_mode))
    return base


def detail_groups(lines: list) -> list[tuple[object, list]]:
    grouped: dict[tuple, list] = {}
    for line in lines:
        key = RENDERER.line_group_key(line)
        grouped.setdefault(key, []).append(line)
    return [(RENDERER.group_invoice_lines(items)[0], items) for items in grouped.values()]


def xlsx_alignment_for_col(col: int, weight_start_col: int, row_kind: str) -> Alignment:
    if row_kind == "header":
        return Alignment(vertical="center", horizontal="center", wrap_text=False)
    if col >= weight_start_col:
        return Alignment(vertical="center", horizontal="right", wrap_text=False)
    if col in {1, 2, 3, 5, 6, 7, 12}:
        return Alignment(vertical="center", horizontal="center", wrap_text=False)
    return Alignment(vertical="center", horizontal="left", wrap_text=False)


def style_row(
    ws,
    row: int,
    col_count: int,
    *,
    weight_start_col: int,
    fill=None,
    font=None,
    border=None,
    row_kind: str = "body",
) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if border:
            cell.border = border
        cell.alignment = xlsx_alignment_for_col(col, weight_start_col, row_kind)


def set_merged_value(ws, row: int, start_col: int, end_col: int, value: str, font=None, alignment=None) -> None:
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment


def render_invoice_detail_xlsx_bytes(invoice) -> tuple[bytes, str]:
    weight_mode = RENDERER.normalise_weight_mode(invoice.weight_mode) or "gross"
    headers = detail_headers(invoice, weight_mode)

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"
    col_count = len(headers)
    weight_start_col = col_count - 1 if weight_mode == "net" else col_count

    header_fill = PatternFill("solid", fgColor="7A0019")
    total_fill = PatternFill("solid", fgColor="A43052")
    group_fill = PatternFill("solid", fgColor="BE737E")
    logo_fills = [
        PatternFill("solid", fgColor="BE737E"),
        PatternFill("solid", fgColor="A43052"),
        PatternFill("solid", fgColor="741C35"),
        PatternFill("solid", fgColor="4D0919"),
    ]
    thin_side = Side(style="thin", color="D5D5D5")
    header_side = Side(style="thin", color="6C0016")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)
    title_font = Font(name="Arial", size=10, bold=True, color="800014")
    small_font = Font(name="Arial", size=8)
    body_font = Font(name="Arial", size=9)
    white_bold = Font(name="Arial", size=8, color="FFFFFF", bold=True)
    bold = Font(name="Arial", size=9, bold=True)
    total_font = Font(name="Arial", size=9, color="FFFFFF", bold=True)
    top_alignment = Alignment(vertical="center", horizontal="left", wrap_text=False, shrink_to_fit=True)

    ws.cell(
        row=1,
        column=1,
        value="Dejar estas columnas ocultas para que no se exporten con el atajo de generar factura agrupada",
    )
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[1].height = 4
    for top_row in range(2, 7):
        ws.row_dimensions[top_row].height = 15
    for row, fill in zip(range(2, 6), logo_fills):
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3).fill = fill
    set_merged_value(ws, 2, 5, 9, "Steel Trade Advisors, S.L.U.", title_font, top_alignment)
    set_merged_value(ws, 3, 5, 9, "C/ Almirante, n\u00ba 22, 5A", small_font, top_alignment)
    set_merged_value(ws, 4, 5, 9, "28004 Madrid, Espa\u00f1a", small_font, top_alignment)
    set_merged_value(ws, 5, 5, 9, "+34 91 068 82 77", small_font, top_alignment)
    set_merged_value(ws, 6, 5, 9, "CIF: B88047790", small_font, top_alignment)
    customer_col = min(10, col_count)
    for offset, line in enumerate(invoice.customer_lines[:5], start=2):
        font = bold if offset == 2 else small_font
        set_merged_value(ws, offset, customer_col, col_count, line, font, top_alignment)

    description = f"Commercial Invoice Detail Number: {invoice.invoice_id}"
    ws.cell(row=8, column=3, value=description)
    ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=max(3, col_count - 1))
    ws.cell(row=8, column=col_count, value=invoice.date_text)
    ws.cell(row=8, column=3).font = Font(name="Arial", size=8, bold=True)
    ws.cell(row=8, column=3).alignment = Alignment(vertical="center", horizontal="left", shrink_to_fit=True)
    ws.cell(row=8, column=col_count).font = Font(name="Arial", size=8)
    ws.cell(row=8, column=col_count).alignment = Alignment(vertical="center", horizontal="right")
    ws.row_dimensions[7].height = 7
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 7

    header_row = 10
    for index, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=index, value=header)
    ws.row_dimensions[header_row].height = 18
    style_row(
        ws,
        header_row,
        col_count,
        weight_start_col=weight_start_col,
        fill=header_fill,
        font=white_bold,
        border=header_border,
        row_kind="header",
    )

    detail_lines = invoice.detail_lines or invoice.lines
    groups = detail_groups(detail_lines)
    total_net = sum(line_net_weight(group_line, weight_mode) for group_line, _ in groups)
    total_gross = sum(line_gross_weight(group_line, weight_mode) for group_line, _ in groups)
    total_row = header_row + 1
    ws.cell(row=total_row, column=3, value="TOTAL")
    if weight_mode == "net":
        ws.cell(row=total_row, column=col_count - 1, value=total_net)
        ws.cell(row=total_row, column=col_count, value=total_gross)
    else:
        ws.cell(row=total_row, column=col_count, value=total_gross)
    ws.row_dimensions[total_row].height = 18
    style_row(
        ws,
        total_row,
        col_count,
        weight_start_col=weight_start_col,
        fill=total_fill,
        font=total_font,
        border=border,
    )

    row = total_row + 1
    for group_line, items in groups:
        item_label = group_line.item_label or group_line.factory_id
        for col, value in enumerate(line_group_values(group_line, weight_mode, invoice), start=1):
            ws.cell(row=row, column=col, value=value)
        ws.row_dimensions[row].height = 18
        style_row(
            ws,
            row,
            col_count,
            weight_start_col=weight_start_col,
            fill=group_fill,
            font=bold,
            border=border,
        )
        row += 1
        for detail_index, line in enumerate(items, start=1):
            for col, value in enumerate(
                line_detail_values(line, weight_mode, invoice, item_label, detail_index), start=1
            ):
                ws.cell(row=row, column=col, value=value)
            ws.row_dimensions[row].height = 17
            style_row(
                ws,
                row,
                col_count,
                weight_start_col=weight_start_col,
                font=body_font,
                border=border,
            )
            row += 1

    last_row = max(total_row, row - 1)
    last_col_letter = get_column_letter(col_count)
    ws.auto_filter.ref = f"C{header_row}:{last_col_letter}{last_row}"
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"C2:{last_col_letter}{last_row}"

    widths = [8, 5, 6, 9.5, 7.5, 8, 8, 11, 12, 13, 15, 18, 11.5, 11.5]
    for index, width in enumerate(widths[:col_count], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].hidden = True
    for row_cells in ws.iter_rows(min_row=total_row, max_row=last_row):
        for cell in row_cells:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000" if cell.column >= weight_start_col else "0"
    ws.freeze_panes = "C12"
    ws.sheet_view.showGridLines = False
    registry_footer = f'&"Arial"&6&K888888{RENDERER.REGISTRY_LINE}'
    ws.oddFooter.center.text = registry_footer
    ws.evenFooter.center.text = registry_footer
    ws.firstFooter.center.text = registry_footer
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.45, header=0.1, footer=0.2)
    ws.sheet_view.zoomScale = 90

    output = BytesIO()
    wb.save(output)
    filename = f"{RENDERER.safe_filename(invoice.invoice_id)}-detalle-existencias.xlsx"
    return output.getvalue(), filename


def monthly_close_number(value) -> float | None:
    formula_error = monthly_close_formula_error(value)
    if formula_error:
        raise ValueError(f"Airtable formula error: {formula_error}")
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"Invalid monthly close number: {value!r}") from exc


def monthly_close_formula_error(value: object) -> str | None:
    if isinstance(value, dict):
        marker = value.get("error") or value.get("specialValue")
        if marker:
            return str(marker)
        for nested in value.values():
            problem = monthly_close_formula_error(nested)
            if problem:
                return problem
        return None
    if isinstance(value, list):
        for nested in value:
            problem = monthly_close_formula_error(nested)
            if problem:
                return problem
        return None
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("{") or text.startswith("["):
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError:
                parsed = None
            if parsed is not None:
                return monthly_close_formula_error(parsed)
        match = re.search(r"#(?:ERROR|DIV/0|VALUE|REF|NAME|N/A|NUM|NULL)!?", text, re.I)
        if match:
            return match.group(0)
    return None


def monthly_close_sheet_name(value: object, index: int, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", str(value or "Packing List"))
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Packing List"
    prefix = f"{index:02d} "
    candidate = f"{prefix}{cleaned}"[:31].rstrip()
    suffix = 2
    while candidate.casefold() in used:
        marker = f" ({suffix})"
        candidate = f"{prefix}{cleaned}"[: 31 - len(marker)].rstrip() + marker
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def monthly_close_formula_sheet_name(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def monthly_close_number_format(field_name: str) -> str:
    if field_name in MONTHLY_CLOSE_WEIGHT_FIELDS:
        return "#,##0.000"
    if field_name in MONTHLY_CLOSE_PRICE_FIELDS:
        return '#,##0.00 "€/MT"'
    if field_name in MONTHLY_CLOSE_NUMERIC_FIELDS:
        return '#,##0.00 "€"'
    return "General"


def style_monthly_close_header(ws, row: int, col_count: int) -> None:
    fill = PatternFill("solid", fgColor="7A0019")
    side = Side(style="thin", color="630014")
    border = Border(bottom=side)
    font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 34


def style_monthly_close_total(ws, row: int, col_count: int) -> None:
    fill = PatternFill("solid", fgColor="EAD5DB")
    side = Side(style="medium", color="7A0019")
    font = Font(name="Arial", size=9, bold=True, color="4D0919")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = Border(top=side)
        cell.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
    ws.row_dimensions[row].height = 20


def render_monthly_close_xlsx_bytes(payload: dict) -> tuple[bytes, str, dict]:
    close_date = str(payload.get("closeDate") or "").strip()
    try:
        date.fromisoformat(close_date)
    except ValueError as exc:
        raise ValueError("closeDate must be a valid ISO date (YYYY-MM-DD).") from exc

    packing_lists = payload.get("packingLists")
    if not isinstance(packing_lists, list) or not packing_lists:
        raise ValueError("packingLists must contain at least one Packing List.")
    if len(packing_lists) > 250:
        raise ValueError("Monthly close is limited to 250 Packing Lists per workbook.")

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    used_sheet_names = {"resumen"}
    detail_sheets: list[dict] = []
    total_existences = 0
    title_font = Font(name="Arial", size=15, bold=True, color="7A0019")
    subtitle_font = Font(name="Arial", size=10, color="595959")
    body_font = Font(name="Arial", size=9, color="202020")
    alternate_fill = PatternFill("solid", fgColor="F8F2F4")
    thin_bottom = Side(style="thin", color="E3D7DB")

    for index, packing_list in enumerate(packing_lists, start=1):
        if not isinstance(packing_list, dict):
            raise ValueError(f"Packing List {index} must be an object.")
        packing_list_name = str(
            packing_list.get("name") or packing_list.get("recordId") or f"Packing List {index}"
        ).strip()
        rows = packing_list.get("existencias")
        if not isinstance(rows, list) or not rows:
            raise ValueError(f"Packing List {packing_list_name} has no existencias.")
        total_existences += len(rows)

        sheet_name = monthly_close_sheet_name(packing_list_name, index, used_sheet_names)
        ws = wb.create_sheet(sheet_name)
        col_count = len(MONTHLY_CLOSE_FIELDS)
        last_col = get_column_letter(col_count)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws.cell(row=1, column=1, value=f"Cierre mensual · {packing_list_name}")
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws.cell(row=2, column=1, value=f"Fecha de cierre: {close_date}")
        ws.cell(row=2, column=1).font = subtitle_font
        ws.row_dimensions[2].height = 18

        header_row = 4
        data_start_row = 5
        for col, field_name in enumerate(MONTHLY_CLOSE_FIELDS, start=1):
            ws.cell(row=header_row, column=col, value=field_name)
        style_monthly_close_header(ws, header_row, col_count)

        for row_offset, item in enumerate(rows):
            if not isinstance(item, dict):
                raise ValueError(f"Existencia {row_offset + 1} in {packing_list_name} must be an object.")
            fields = item.get("fields")
            if not isinstance(fields, dict):
                raise ValueError(
                    f"Existencia {row_offset + 1} in {packing_list_name} is missing fields."
                )
            excel_row = data_start_row + row_offset
            for col, field_name in enumerate(MONTHLY_CLOSE_FIELDS, start=1):
                raw_value = fields.get(field_name)
                try:
                    value = (
                        monthly_close_number(raw_value)
                        if field_name in MONTHLY_CLOSE_NUMERIC_FIELDS
                        else raw_value
                    )
                except ValueError as exc:
                    existence_name = str(
                        fields.get("ID Fábrica")
                        or item.get("recordId")
                        or f"Existencia {row_offset + 1}"
                    ).strip()
                    raise ValueError(
                        f"{packing_list_name} / {existence_name} / {field_name}: {exc}"
                    ) from exc
                cell = ws.cell(row=excel_row, column=col, value=value)
                cell.font = body_font
                cell.alignment = Alignment(
                    horizontal="right" if field_name in MONTHLY_CLOSE_NUMERIC_FIELDS else "left",
                    vertical="center",
                )
                if field_name in MONTHLY_CLOSE_NUMERIC_FIELDS:
                    cell.number_format = monthly_close_number_format(field_name)
                cell.border = Border(bottom=thin_bottom)
                if row_offset % 2:
                    cell.fill = alternate_fill
            ws.row_dimensions[excel_row].height = 18

        data_end_row = data_start_row + len(rows) - 1
        total_row = data_end_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        value_total_col = MONTHLY_CLOSE_FIELDS.index("Valor total") + 1
        net_col = MONTHLY_CLOSE_FIELDS.index("Neto") + 1
        for col, field_name in enumerate(MONTHLY_CLOSE_FIELDS[1:], start=2):
            col_letter = get_column_letter(col)
            if field_name == "Precio total":
                value_letter = get_column_letter(value_total_col)
                net_letter = get_column_letter(net_col)
                formula = f"=IFERROR({value_letter}{total_row}/{net_letter}{total_row},0)"
            else:
                formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
            ws.cell(row=total_row, column=col, value=formula)
            ws.cell(row=total_row, column=col).number_format = monthly_close_number_format(field_name)
        style_monthly_close_total(ws, total_row, col_count)

        ws.auto_filter.ref = f"A{header_row}:{last_col}{data_end_row}"
        ws.freeze_panes = f"A{data_start_row}"
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 75
        ws.print_title_rows = f"1:{header_row}"
        ws.print_area = f"A1:{last_col}{total_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.45, header=0.15, footer=0.2)
        ws.oddFooter.center.text = f'&"Arial"&7&K777777Cierre mensual {close_date}'
        ws.evenFooter.center.text = ws.oddFooter.center.text

        for col, field_name in enumerate(MONTHLY_CLOSE_FIELDS, start=1):
            if field_name == "ID Fábrica":
                width = 20
            elif field_name in MONTHLY_CLOSE_WEIGHT_FIELDS:
                width = 13
            elif field_name in MONTHLY_CLOSE_PRICE_FIELDS:
                width = 16
            elif field_name.startswith("Movimientos de compra"):
                width = 28
            elif len(field_name) > 24:
                width = 23
            else:
                width = 18
            ws.column_dimensions[get_column_letter(col)].width = width

        detail_sheets.append(
            {
                "name": sheet_name,
                "packingList": packing_list_name,
                "totalRow": total_row,
                "dataStartRow": data_start_row,
                "dataEndRow": data_end_row,
                "recordId": packing_list.get("recordId"),
            }
        )

    summary_headers = ["Packing List", "Existencias", *MONTHLY_CLOSE_FIELDS[1:]]
    summary = wb.create_sheet("Resumen", 0)
    summary_col_count = len(summary_headers)
    summary_last_col = get_column_letter(summary_col_count)
    summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=summary_col_count)
    summary.cell(row=1, column=1, value="Cierre mensual de existencias contables")
    summary.cell(row=1, column=1).font = title_font
    summary.row_dimensions[1].height = 26
    summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=summary_col_count)
    summary.cell(row=2, column=1, value=f"Fecha de cierre: {close_date}")
    summary.cell(row=2, column=1).font = subtitle_font

    summary_header_row = 4
    summary_data_start = 5
    for col, header in enumerate(summary_headers, start=1):
        summary.cell(row=summary_header_row, column=col, value=header)
    style_monthly_close_header(summary, summary_header_row, summary_col_count)

    for row_offset, detail in enumerate(detail_sheets):
        row = summary_data_start + row_offset
        quoted_sheet = monthly_close_formula_sheet_name(detail["name"])
        summary.cell(row=row, column=1, value=detail["packingList"])
        summary.cell(row=row, column=1).hyperlink = f"#{quoted_sheet}!A1"
        summary.cell(row=row, column=1).style = "Hyperlink"
        summary.cell(
            row=row,
            column=2,
            value=f"=ROWS({quoted_sheet}!A{detail['dataStartRow']}:A{detail['dataEndRow']})",
        )
        for detail_col, field_name in enumerate(MONTHLY_CLOSE_FIELDS[1:], start=2):
            summary_col = detail_col + 1
            if field_name == "Precio total":
                value_summary_col = 2 + MONTHLY_CLOSE_FIELDS.index("Valor total")
                net_summary_col = 2 + MONTHLY_CLOSE_FIELDS.index("Neto")
                formula = (
                    f"=IFERROR({get_column_letter(value_summary_col)}{row}/"
                    f"{get_column_letter(net_summary_col)}{row},0)"
                )
            else:
                formula = f"={quoted_sheet}!{get_column_letter(detail_col)}{detail['totalRow']}"
            cell = summary.cell(row=row, column=summary_col, value=formula)
            cell.number_format = monthly_close_number_format(field_name)
        for col in range(1, summary_col_count + 1):
            cell = summary.cell(row=row, column=col)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
            cell.border = Border(bottom=thin_bottom)
            if row_offset % 2:
                cell.fill = alternate_fill
        summary.row_dimensions[row].height = 18

    summary_data_end = summary_data_start + len(detail_sheets) - 1
    summary_total_row = summary_data_end + 1
    summary.cell(row=summary_total_row, column=1, value="TOTAL GENERAL")
    for col in range(2, summary_col_count + 1):
        field_name = summary_headers[col - 1]
        col_letter = get_column_letter(col)
        if field_name == "Precio total":
            value_col = summary_headers.index("Valor total") + 1
            net_col_summary = summary_headers.index("Neto") + 1
            formula = (
                f"=IFERROR({get_column_letter(value_col)}{summary_total_row}/"
                f"{get_column_letter(net_col_summary)}{summary_total_row},0)"
            )
        else:
            formula = f"=SUM({col_letter}{summary_data_start}:{col_letter}{summary_data_end})"
        summary.cell(row=summary_total_row, column=col, value=formula)
        if field_name in MONTHLY_CLOSE_NUMERIC_FIELDS:
            summary.cell(row=summary_total_row, column=col).number_format = monthly_close_number_format(field_name)
        elif field_name == "Existencias":
            summary.cell(row=summary_total_row, column=col).number_format = "#,##0"
    style_monthly_close_total(summary, summary_total_row, summary_col_count)

    summary.auto_filter.ref = f"A{summary_header_row}:{summary_last_col}{summary_data_end}"
    summary.freeze_panes = f"A{summary_data_start}"
    summary.sheet_view.showGridLines = False
    summary.sheet_view.zoomScale = 75
    summary.print_title_rows = f"1:{summary_header_row}"
    summary.print_area = f"A1:{summary_last_col}{summary_total_row}"
    summary.page_setup.orientation = "landscape"
    summary.page_setup.fitToWidth = 1
    summary.page_setup.fitToHeight = 0
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.45, header=0.15, footer=0.2)
    summary.oddFooter.center.text = f'&"Arial"&7&K777777Cierre mensual {close_date}'
    summary.evenFooter.center.text = summary.oddFooter.center.text
    for col, field_name in enumerate(summary_headers, start=1):
        if field_name == "Packing List":
            width = 24
        elif field_name == "Existencias":
            width = 12
        elif field_name in MONTHLY_CLOSE_WEIGHT_FIELDS:
            width = 13
        elif field_name in MONTHLY_CLOSE_PRICE_FIELDS:
            width = 16
        elif field_name.startswith("Movimientos de compra"):
            width = 28
        elif len(field_name) > 24:
            width = 23
        else:
            width = 18
        summary.column_dimensions[get_column_letter(col)].width = width

    output = BytesIO()
    wb.save(output)
    filename = f"cierre-mensual-{close_date}.xlsx"
    metadata = {
        "closeDate": close_date,
        "packingLists": len(detail_sheets),
        "existencias": total_existences,
        "sheets": ["Resumen", *[item["name"] for item in detail_sheets]],
    }
    return output.getvalue(), filename, metadata


def clean_file_cache(config: dict) -> None:
    now = time.time()
    ttl = int(config["cache_ttl_seconds"])
    expired = [
        key
        for key, item in FILE_CACHE.items()
        if now - float(item.get("created_at", 0)) > ttl
    ]
    for key in expired:
        FILE_CACHE.pop(key, None)


def cache_file(file_bytes: bytes, filename: str, content_type: str, config: dict) -> str:
    clean_file_cache(config)
    suffix = Path(filename).suffix or ".bin"
    key = f"{secrets.token_urlsafe(24)}{suffix}"
    FILE_CACHE[key] = {
        "bytes": file_bytes,
        "filename": filename,
        "content_type": content_type,
        "created_at": time.time(),
    }
    return key


def public_base_url(handler: BaseHTTPRequestHandler, config: dict) -> str:
    configured = str(config.get("public_base_url") or "").strip().rstrip("/")
    if configured:
        return configured
    proto = handler.headers.get("X-Forwarded-Proto") or "http"
    host = handler.headers.get("Host") or f"{DEFAULT_HOST}:{DEFAULT_PORT}"
    return f"{proto}://{host}"


def upload_attachment(config: dict, record_id: str, pdf_bytes: bytes, filename: str) -> dict:
    token = str(config.get("airtable_token") or "")
    field_id = str(config.get("attachment_field_id") or "")
    if not token:
        raise ValueError("Upload requires AIRTABLE_TOKEN or AIRTABLE_PAT.")
    if not field_id:
        raise ValueError("Upload requires STA_INVOICE_ATTACHMENT_FIELD_ID.")
    if not record_id:
        raise ValueError("Upload requires an Airtable record ID.")

    url = (
        f"https://content.airtable.com/v0/{config['base_id']}/"
        f"{record_id}/{field_id}/uploadAttachment"
    )
    payload = {
        "contentType": "application/pdf",
        "filename": filename,
        "file": base64.b64encode(pdf_bytes).decode("ascii"),
    }
    return RENDERER.airtable_request_json(url, token, method="POST", payload=payload)


class Handler(BaseHTTPRequestHandler):
    server_version = "STAInvoicePDF/0.3"

    def do_OPTIONS(self) -> None:  # noqa: N802 - stdlib API
        self.send_response(HTTPStatus.NO_CONTENT)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Authorization, Content-Type, X-STA-Invoice-Token")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.end_headers()

    def do_GET(self) -> None:  # noqa: N802 - stdlib API
        config = service_config()
        parsed = urlparse(self.path)
        if parsed.path == "/health":
            json_response(
                self,
                HTTPStatus.OK,
                {
                    "ok": True,
                    "serviceVersion": self.server_version,
                    "baseId": config["base_id"],
                    "hasAirtableToken": bool(config["airtable_token"]),
                    "hasAttachmentField": bool(config["attachment_field_id"]),
                    "publicBaseUrl": config["public_base_url"],
                    "cacheTtlSeconds": config["cache_ttl_seconds"],
                    "monthlyCloseEndpoint": "/airtable/monthly-close/render",
                },
            )
            return
        if parsed.path == "/airtable/reassignment-map":
            if not REASSIGNMENT_MAP_HTML.exists():
                text_response(self, HTTPStatus.NOT_FOUND, "Reassignment map not found")
                return
            html_response(self, HTTPStatus.OK, REASSIGNMENT_MAP_HTML.read_bytes())
            return
        if parsed.path.startswith("/files/"):
            clean_file_cache(config)
            key = parsed.path.removeprefix("/files/")
            item = FILE_CACHE.get(key)
            if not item:
                text_response(self, HTTPStatus.NOT_FOUND, "File not found or expired")
                return
            file_bytes = item["bytes"]
            filename = item["filename"]
            content_type = item.get("content_type") or "application/octet-stream"
            disposition = "inline" if content_type == "application/pdf" else "attachment"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Disposition", f'{disposition}; filename="{filename}"')
            self.send_header("Content-Length", str(len(file_bytes)))
            self.end_headers()
            self.wfile.write(file_bytes)
            return
        if parsed.path != "/invoice.pdf":
            text_response(self, HTTPStatus.NOT_FOUND, "Not found")
            return
        if not is_authorized(self, config):
            text_response(self, HTTPStatus.UNAUTHORIZED, "Unauthorized")
            return
        try:
            query = parse_qs(parsed.query)
            payload = {key: values[0] for key, values in query.items()}
            invoice = build_invoice(payload, config)
            pdf_bytes, filename = render_invoice_pdf_bytes(invoice, int(config["max_lines"]))
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Content-Disposition", f'inline; filename="{filename}"')
            self.send_header("Content-Length", str(len(pdf_bytes)))
            self.end_headers()
            self.wfile.write(pdf_bytes)
        except Exception as exc:  # pragma: no cover - HTTP guard
            json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})

    def do_POST(self) -> None:  # noqa: N802 - stdlib API
        config = service_config()
        parsed = urlparse(self.path)
        if parsed.path not in {
            "/airtable/invoices/generate",
            "/airtable/invoices/render",
            "/airtable/monthly-close/render",
        }:
            text_response(self, HTTPStatus.NOT_FOUND, "Not found")
            return
        if not is_authorized(self, config):
            text_response(self, HTTPStatus.UNAUTHORIZED, "Unauthorized")
            return
        try:
            payload = read_json_body(self)
            if parsed.path == "/airtable/monthly-close/render":
                xlsx_bytes, filename, metadata = render_monthly_close_xlsx_bytes(payload)
                key = cache_file(xlsx_bytes, filename, MONTHLY_CLOSE_CONTENT_TYPE, config)
                attachment = {
                    "url": f"{public_base_url(self, config)}/files/{key}",
                    "filename": filename,
                    "contentType": MONTHLY_CLOSE_CONTENT_TYPE,
                    "bytes": len(xlsx_bytes),
                }
                json_response(
                    self,
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "filename": filename,
                        "bytes": len(xlsx_bytes),
                        "fileUrl": attachment["url"],
                        "attachments": [attachment],
                        "metadata": metadata,
                        "expiresInSeconds": config["cache_ttl_seconds"],
                    },
                )
                return
            if parsed.path == "/airtable/invoices/render":
                invoice = build_invoice_from_payload(payload)
            else:
                invoice = build_invoice(payload, config)
            pdf_bytes, filename = render_invoice_pdf_bytes(invoice, int(config["max_lines"]))
            if parsed.path == "/airtable/invoices/render":
                rendered_files = [
                    {
                        "bytes": pdf_bytes,
                        "filename": filename,
                        "contentType": "application/pdf",
                    }
                ]
                if invoice.mode == "grouped":
                    xlsx_bytes, xlsx_filename = render_invoice_detail_xlsx_bytes(invoice)
                    rendered_files.append(
                        {
                            "bytes": xlsx_bytes,
                            "filename": xlsx_filename,
                            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        }
                    )
                attachments = []
                for item in rendered_files:
                    key = cache_file(item["bytes"], item["filename"], item["contentType"], config)
                    attachments.append(
                        {
                            "url": f"{public_base_url(self, config)}/files/{key}",
                            "filename": item["filename"],
                            "contentType": item["contentType"],
                            "bytes": len(item["bytes"]),
                        }
                    )
                file_url = attachments[0]["url"]
                json_response(
                    self,
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "invoice": invoice.invoice_id,
                        "recordId": invoice.record_id,
                        "filename": filename,
                        "bytes": len(pdf_bytes),
                        "fileUrl": file_url,
                        "attachments": attachments,
                        "expiresInSeconds": config["cache_ttl_seconds"],
                    },
                )
                return
            should_upload = bool(payload.get("upload", config["default_upload"]))
            upload_result = None
            if should_upload:
                upload_result = upload_attachment(
                    config=config,
                    record_id=invoice.record_id,
                    pdf_bytes=pdf_bytes,
                    filename=filename,
                )
            json_response(
                self,
                HTTPStatus.OK,
                {
                    "ok": True,
                    "invoice": invoice.invoice_id,
                    "recordId": invoice.record_id,
                    "filename": filename,
                    "bytes": len(pdf_bytes),
                    "uploaded": upload_result is not None,
                    "uploadResult": upload_result,
                },
            )
        except Exception as exc:  # pragma: no cover - HTTP guard
            json_response(self, HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})

    def log_message(self, format: str, *args) -> None:  # noqa: A002 - stdlib API
        print(f"{self.address_string()} - {format % args}")


def main() -> int:
    default_host = "0.0.0.0" if os.environ.get("PORT") else DEFAULT_HOST
    host = os.environ.get("STA_INVOICE_HOST", default_host)
    port = int(os.environ.get("PORT") or os.environ.get("STA_INVOICE_PORT", str(DEFAULT_PORT)))
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"STA invoice PDF service listening on http://{host}:{port}")
    print("GET  /health")
    print("GET  /airtable/reassignment-map")
    print("GET  /invoice.pdf?invoice=2025-192")
    print("POST /airtable/invoices/generate")
    print("POST /airtable/invoices/render")
    print("POST /airtable/monthly-close/render")
    server.serve_forever()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
